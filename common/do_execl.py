#!/usr/bin/env python
# -*- coding:utf-8-*-
#@author:蜜蜜
#@file: do_execl.py 
#@time: 2018/12/18 
#@email:1402686685@qq.com
import openpyxl
import json
import os
from common import contants
from common.request import Request


class Case:  #测试类的封装
    def __init__(self):#初始化方法,把列写出来
        self.case_id=None
        self.url=None
        self.data=None
        self.title=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None


class DoExecl:#execl操作的封装

    def  __init__(self,file_name):#DoExecl初始化函数
        try:
            self.file_name=file_name #把属性传给对象

            #打开一个execl文件，返回一个workbook对象实例，把它定义为
            #DoExecl的属性，以便在这个类的其他地方调用,以前直接默认为filename存在
            self.workbook=openpyxl.load_workbook(filename=file_name)#, #把属性传给对象
            # self.workbook = openpyxl.load_workbook(file_name)  # , #把属性传给对象
            # self.workbook = openpyxl.load_workbook(self.file_name)  # , #把属性传给对象
        except FileNotFoundError as e:#文件为找到异常处理
            print('{}not found,please check file path'.format(e))
            raise e #抛出异常


    def get_cases(self,sheet_name):#根据sheet名称，获取在这个sheet里面的所有测试用例数据
        sheet=self.workbook[sheet_name]#根据sheet名称获取sheet对象实例
        max_row=sheet.max_row#获取sheet最大行数
        cases=[]#定义一个列表，用来存放即将要放进去的测试用例

        # for r in range(2,sheet.max_row):
        for r in range (2,max_row+1):#左闭右开，从第二行开始取值,最大行+1
            case=Case()#实例化一个case对象，用来存放测试数据
            case.id=sheet.cell(row=r,column=1).value#取第r行，第1格的值，对象调用id属性
            case.title = sheet.cell(row=r, column=2).value  # 取第r行，第4格的值，对象调用title属性
            case.url = sheet.cell(row=r, column=3).value  # 取第r行，第2格的值，对象调用url属性
            case.data = sheet.cell(row=r, column=4).value  # 取第r行，第3格的值，对象调用data属性
            case.method = sheet.cell(row=r, column=5).value  # 取第r行，第5格的值，对象调用method属性
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6格的值，对象调用expected属性
            cases.append(case)#将case放到cases列表里面
        return cases#for循环结束返回cases列表

    def get_sheet_names(self):#定义一个方法，获取所有的sheet
        return self.workbook.sheetnames

    # 根据sheet_name定位sheet，根据case_id定位行，取到当前行的actual单元
    def write_by_case_id(self,sheet_name,case_id,actual,result):
        sheet=self.workbook[sheet_name]#根据sheet名称获取sheet对象实例
        max_row=sheet.max_row#获取sheet最大行数
        for r in range(2,max_row+1):#for 循环，从第二行开始遍历
            case_id_r= sheet.cell(r,1).value#取第r行第一列，即case_id
            if case_id_r==case_id:  #判断execl里面取到的当前行的case_id是否等于传进来的case_id
                sheet.cell(r,7).value=actual#写入传进来的actual到当前的actual单元格
                sheet.cell(r, 8).value = result  # 写入传进来的actual到当前的result单元格
                self.workbook.save(filename=self.file_name)
                break

if __name__ == '__main__':
    print('开始调试')
    #测试一下DcExecl类

    do_execl=DoExecl(os.path.join(contants.datas_dir,'cases.xlsx'))#实例化为一个DoExecl对象
    sheet_names=do_execl.get_sheet_names()#获取execl文件中全部sheet名称
    # cases=do_execl.get_cases('login')
    # print(cases)
    # for case in cases:
    #     print(case.id)
    #     print(case.url)
    print("sheet 名称列表：",sheet_names)
    cases_list=['login','register']#定义一个执行测试用例的列表
    #根据sheet_names分别获取到全部用例来执行
    for sheet_name in sheet_names:
        #获取login的所有测试用例，返回一个casts列表，列表里面是多个case类实例，每个实例的属性对应execl表中每列的值
        if sheet_name in cases_list:#如果当前的sheet_name不在可执行放case__name就不执行
            cases=do_execl.get_cases(sheet_name)
            print(sheet_name+'测试用例个数：',len(cases))
            for case in cases:#遍历测试用例列表，每进for一次，就取一个case实例
                print("case信息：", case.__dict__)  # 打印case信息
                # print(type(case.data))
                data = eval(case.data)  # Excel里面取到data是一个字符串，使用eval函数将字符串转换成字典
                # print(type(data))
                resp = Request(method=case.method, url=case.url, data=data)  # 通过封装的Request类来完成接口的调用
                print('status_code:', resp.get_status_code())  # 打印响应码
                resp_dict = resp.get_json()  # 获取请求响应，字典
                # 通过json.dumps函数将字典转换成格式化后的字符串
                resp_text = json.dumps(resp_dict, ensure_ascii=False, indent=4)
                print('response: ', resp_text)  # 打印响应
                # 判断接口响应是否和Excel里面expected的值是否一致
                if case.expected == resp.get_text():
                    print("result : PASS")
                    do_execl.write_by_case_id(sheet_name=sheet_name,case_id=case.id,actual=resp.get_text(),result='PASS')#期望结果与实际结果一致，就写入PASS到result这个单元格
                else:
                    print("result : FAIL")
                    do_execl.write_by_case_id(sheet_name=sheet_name,case_id=case.id,actual=resp.get_text(),result='FAIL')#期望结果与实际结果一致，就写入FAIL到result这个单元格



